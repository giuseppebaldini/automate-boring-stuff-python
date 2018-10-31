# Chapter 12 - Working with Excel with OpenPyXL

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)

# Getting sheets from the Workbook
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet3')
type(sheet)
sheet.title
activeSheet = wb.active

# Getting cells from the Sheets
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value

# Assing cell value to variable
c = sheet['B1']

# Alternatively
sheet.cell(row = 1, column = 2).value

# Finding the size of the sheet (always returning int)
sheet.max_row
sheet.max_column

# Selecting areas of the spreadsheet
tuple(sheet['A1':'C3'])
